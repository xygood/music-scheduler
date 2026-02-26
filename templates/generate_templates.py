#!/usr/bin/env python3
"""生成音乐学校排课系统数据导入模板"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 设置中文字体
try:
    from openpyxl.styles import Font
    font = Font(name='SimHei', size=11)
except:
    font = Font(size=11)

# 样式定义
header_fill = PatternFill(start_color="6B5B95", end_color="6B5B95", fill_type="solid")
header_font = Font(name='SimHei', size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(ws, widths):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_students_template():
    """创建学生导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学生"

    # 表头
    headers = ["学号", "姓名", "乐器", "年级"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 示例数据
    sample_data = [
        ["S0001", "张三", "钢琴", "一年级"],
        ["S0002", "李四", "声乐", "二年级"],
        ["S0003", "王五", "小提琴", "研究生"],
        ["S0004", "赵六", "古筝", "三年级"],
        ["S0005", "陈七", "笛子", "一年级"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    set_column_widths(ws, [12, 12, 12, 12])

    # 添加说明
    ws.cell(row=8, column=1, value="说明：学号可留空自动生成，专业支持：钢琴、声乐、双排键、小提琴、古筝、笛子、古琴、葫芦丝、萨克斯")

    return wb

def create_courses_template():
    """创建课程导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "课程"

    # 表头
    headers = ["课程名称", "课程类型", "学生姓名", "课时长度", "每周次数"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 示例数据
    sample_data = [
        ["钢琴基础练习", "钢琴", "张三", "30", "2"],
        ["声乐演唱技巧", "声乐", "李四", "45", "1"],
        ["小提琴独奏", "器乐", "王五", "60", "2"],
        ["古筝基础入门", "器乐", "赵六", "30", "1"],
        ["双排键演奏", "器乐", "陈七", "45", "2"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    set_column_widths(ws, [18, 12, 12, 12, 12])

    # 添加说明
    ws.cell(row=8, column=1, value="说明：课程类型支持：钢琴、声乐、器乐")

    return wb

def create_rooms_template():
    """创建教室导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "教室"

    # 表头
    headers = ["教室名称", "教室类型", "容量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 示例数据
    sample_data = [
        ["101琴房", "琴房", "1"],
        ["102琴房", "琴房", "1"],
        ["201教室", "教室", "30"],
        ["A101大教室", "大教室", "50"],
        ["301排练厅", "排练厅", "50"],
        ["202琴房", "琴房", "1"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    set_column_widths(ws, [15, 12, 10])

    # 添加说明
    ws.cell(row=9, column=1, value="说明：教室类型支持：琴房、教室、大教室、排练厅")

    return wb

def main():
    """主函数"""
    output_dir = "templates"
    os.makedirs(output_dir, exist_ok=True)

    # 生成各个模板
    students_wb = create_students_template()
    students_wb.save(os.path.join(output_dir, "学生导入模板.xlsx"))

    courses_wb = create_courses_template()
    courses_wb.save(os.path.join(output_dir, "课程导入模板.xlsx"))

    rooms_wb = create_rooms_template()
    rooms_wb.save(os.path.join(output_dir, "教室导入模板.xlsx"))

    print("✅ 模板文件生成完成！")
    print(f"📁 文件位置: {output_dir}/")
    print("  - 学生导入模板.xlsx")
    print("  - 课程导入模板.xlsx")
    print("  - 教室导入模板.xlsx")

if __name__ == "__main__":
    main()
